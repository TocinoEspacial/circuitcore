from typing import Self
from django import forms
from django.contrib.auth import login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.views import View
from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.forms import modelformset_factory
from django.contrib.auth.decorators import login_required
from decimal import Decimal
from django.contrib.auth.models import Group
from django.contrib import messages
from openpyxl import Workbook

from .models import Cotizacion, CotizacionItem
from .forms import (
    CotizacionForm,
    CotizacionItemForm,
    CotizacionItemIngenieroForm,
    CotizacionRevisarForm,
    CustomUserCreationForm
)

# ---------- Autenticación y registro ----------
class LoginView(View):
    def get(self, request):
        if request.user.is_authenticated:
            return redirect(self.get_redirect_url(request.user))
        form = AuthenticationForm()
        return render(request, 'registration/login.html', {'form': form})

    def post(self, request):
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect(self.get_redirect_url(user))
        return render(request, 'registration/login.html', {'form': form})

    def get_redirect_url(self, user):
        if user.is_superuser:
            return '/admin/'
        elif user.groups.filter(id=1).exists():
            return 'home'
        elif user.groups.filter(id= 3).exists():
            return 'home'
        elif user.groups.filter(id=2).exists():
            return 'home'
        return 'home'

def register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Asignar grupo Cliente por defecto
            cliente_group = Group.objects.get(name='Cliente') 
            user.groups.add(cliente_group)
            login(request, user)
            return redirect('home')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/register.html', {'form': form})

def home(request):
    return render(request, 'registration/home.html')

# ---------- Vistas por rol ----------
@login_required
def ingeniero_home(request):
    if not request.user.groups.filter(name='Ingeniero').exists():
        return redirect('home')
    return render(request, 'roles/ingeniero.html')

@login_required
def constructor_home(request):
    if not request.user.groups.filter(name='Constructor').exists():
        return redirect('home')
    return render(request, 'roles/constructor.html')

@login_required
def cliente_home(request):
    if not request.user.groups.filter(name='Cliente').exists():
        return redirect('home')
    return render(request, 'roles/cliente.html')

# ---------- Utilidades ----------
def es_ingeniero(user):
    return user.groups.filter(name='ingeniero').exists()

# ---------- Cotizaciones ----------
@login_required
def cotizaciones_pendientes(request):
    if es_ingeniero(request.user):
        cotizaciones = Cotizacion.objects.filter(estado='BORRADOR')
    else:
        cotizaciones = Cotizacion.objects.filter(cliente=request.user, estado='pendiente')
    
    return render(request, 'cotizacion/lista_pendientes.html', {
        'cotizaciones': cotizaciones
    })

from django.shortcuts import render, get_object_or_404, redirect
from django.forms import inlineformset_factory
from django.contrib.auth.decorators import login_required, permission_required
from decimal import Decimal

from .models import Cotizacion, CotizacionItem, PerfilUsuario
from .forms import CotizacionForm, CotizacionItemForm

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.forms import inlineformset_factory
from .models import Cotizacion, CotizacionItem
from .forms import CotizacionForm, CotizacionItemForm

@login_required
def revisar_cotizacion(request, cotizacion_id):
    cotizacion = get_object_or_404(Cotizacion, pk=cotizacion_id)
    
    # Formset específico para ingenieros
    CotizacionItemIngenieroFormSet = inlineformset_factory(
        Cotizacion,
        CotizacionItem,
        form=CotizacionItemIngenieroForm,
        extra=0,
        can_delete=False,
        fields=['descripcion', 'cantidad', 'valor_unidad', 'unidad_medida']
    )

    if request.method == 'POST':
        form = CotizacionRevisarForm(request.POST, instance=cotizacion)
        formset = CotizacionItemIngenieroFormSet(
            request.POST, 
            instance=cotizacion,
            queryset=CotizacionItem.objects.filter(cotizacion=cotizacion)
        )

        if form.is_valid() and formset.is_valid():
            # Guardar la cotización primero
            cotizacion = form.save(commit=False)
            cotizacion.ingeniero = request.user
            cotizacion.estado = 'revisado'  # Marcamos como revisado aquí
            
            # Guardar los items con los nuevos valores unitarios
            instances = formset.save(commit=False)
            for instance in instances:
                instance.valor_total = instance.cantidad * instance.valor_unidad
                instance.save()
            
            # Guardar la cotización después de los items
            cotizacion.save()
            
            # Actualizar los totales de la cotización
            cotizacion.actualizar_totales()

            messages.success(request, 'Cotización revisada y actualizada exitosamente!')
            return redirect('cotizaciones_pendientes')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = CotizacionRevisarForm(instance=cotizacion)
        formset = CotizacionItemIngenieroFormSet(
            instance=cotizacion,
            queryset=CotizacionItem.objects.filter(cotizacion=cotizacion)
        )

    context = {
        'cotizacion': cotizacion,
        'form': form,
        'formset': formset,
    }
    return render(request, 'cotizacion/revisar.html', context)

from django.db import transaction
from django.shortcuts import render, redirect
from django.forms import inlineformset_factory
from django.contrib import messages
from decimal import Decimal
from .models import Cotizacion, CotizacionItem
from .forms import CotizacionForm

from django.db import transaction
from django.shortcuts import render, redirect
from django.forms import inlineformset_factory
from django.contrib import messages
from decimal import Decimal
from .models import Cotizacion, CotizacionItem
from .forms import CotizacionForm

from openpyxl import load_workbook
from decimal import Decimal
from django.db import transaction
from django.contrib import messages

@login_required
def solicitar_cotizacion(request):
    CotizacionItemFormset = inlineformset_factory(
        Cotizacion,
        CotizacionItem,
        fields=['descripcion', 'cantidad', 'unidad_medida'],
        extra=1,
        can_delete=True
    )
    
    if request.method == 'POST':
        # Si se envió archivo Excel, procesarlo
        if 'excel_file' in request.FILES and request.FILES['excel_file']:
            excel_file = request.FILES['excel_file']
            try:
                wb = load_workbook(filename=excel_file, data_only=True)
                ws = wb.active
                
                with transaction.atomic():
                    # Aquí asumo que el archivo Excel tiene las columnas: descripción, cantidad, unidad_medida (ajusta según tu archivo)
                    cotizacion = Cotizacion(
                        cliente=request.user,
                        contacto=request.user.get_full_name() or request.user.username,
                        ciudad=Cotizacion.Ciudades.BOGOTA,
                        tipo_pago=Cotizacion.TipoPago.CONTADO,
                        estado=Cotizacion.Estados.BORRADOR,
                        aplica_iva=True,
                    )
                    cotizacion.save()
                    
                    # Empezar a leer desde fila 2 por ejemplo, para saltar encabezado
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        descripcion, cantidad, unidad_medida = row[:3]
                        if descripcion and cantidad:
                            item = CotizacionItem(
                                cotizacion=cotizacion,
                                descripcion=descripcion,
                                cantidad=Decimal(cantidad),
                                unidad_medida=unidad_medida,
                                valor_unidad=Decimal('0.00'),  # Valor por defecto
                            )
                            item.save()
                    
                    cotizacion.actualizar_totales()
                    
                    messages.success(request, "¡Cotización cargada exitosamente desde Excel!")
                    return redirect('solicitar')
            except Exception as e:
                messages.error(request, f"Error al procesar el archivo Excel: {e}")
                # Al mostrar el error, dejamos el formulario normal para que pueda reintentar
                
        else:
            # Procesar formulario normal con formset
            form = CotizacionForm(request.POST)
            formset = CotizacionItemFormset(request.POST, prefix='items')
            
            if form.is_valid() and formset.is_valid():
                try:
                    with transaction.atomic():
                        cotizacion = form.save(commit=False)
                        cotizacion.cliente = request.user
                        cotizacion.save()
                        
                        instances = formset.save(commit=False)
                        for instance in instances:
                            instance.cotizacion = cotizacion
                            instance.valor_unidad = Decimal('0.00')
                            instance.save()
                        
                        cotizacion.actualizar_totales()
                        
                        messages.success(request, "¡Cotización creada exitosamente!")
                        return redirect('solicitar')
                except Exception as e:
                    messages.error(request, f"Error al guardar: {e}")
    else:
        form = CotizacionForm(initial={
            'contacto': request.user.get_full_name() or request.user.username,
            'ciudad': Cotizacion.Ciudades.BOGOTA,
            'tipo_pago': Cotizacion.TipoPago.CONTADO,
            'estado': Cotizacion.Estados.BORRADOR,
            'aplica_iva': True,
        })
        formset = CotizacionItemFormset(prefix='items', queryset=CotizacionItem.objects.none())
    
    return render(request, 'cotizacion/solicitar.html', {
        'form': form,
        'formset': formset
    })

def actualizar_totales(self, cotizacion):
    """Función auxiliar para actualizar los totales de la cotización"""
    from django.db.models import Sum, F
    
    # Calcula el subtotal sumando (cantidad * valor_unidad) de todos los ítems
    resultado = cotizacion.items.aggregate(
        subtotal=Sum(F('cantidad') * F('valor_unidad')))
    
    cotizacion.subtotal = resultado['subtotal'] or Decimal('0.00')
    cotizacion.iva = cotizacion.subtotal * Decimal('0.19') if cotizacion.aplica_iva else Decimal('0.00')
    cotizacion.total = cotizacion.subtotal + cotizacion.iva - cotizacion.descuento
    cotizacion.save(update_fields=['subtotal', 'iva', 'total'])
    
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Cotizacion, PerfilUsuario
from .forms import PerfilForm, UserForm

@login_required
def editar_perfil(request):
    user = request.user
    perfil, created = PerfilUsuario.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=user)
        perfil_form = PerfilForm(request.POST, request.FILES, instance=perfil)
        
        if user_form.is_valid() and perfil_form.is_valid():
            user_form.save()
            perfil_form.save()
            
            # Redirigir según el tipo de usuario
            if user.groups.filter(name='Ingeniero').exists():
                return redirect('perfil_ingeniero')
            else:
                return redirect('perfil_cliente')
    else:
        user_form = UserForm(instance=user)
        perfil_form = PerfilForm(instance=perfil)
    
    return render(request, 'ingeniero/editar_perfil.html', {
        'user_form': user_form,
        'perfil_form': perfil_form
    })

@login_required
def perfil_ingeniero(request):
    user = request.user
    
    if not user.groups.filter(name='Ingeniero').exists():
        return redirect('home')
    
    perfil, created = PerfilUsuario.objects.get_or_create(user=user)
    
    # Obtener cotizaciones donde el usuario actual es el ingeniero asignado
    cotizaciones = Cotizacion.objects.filter(ingeniero=user).order_by('-fecha')
    
    context = {
        'cotizaciones': cotizaciones,
        'user': user,
        'perfil': perfil,
        'rol': 'Ingeniero',
        'avatar_url': perfil.avatar.url if perfil.avatar else '/static/images/default-avatar.png'
    }
    
    return render(request, 'ingeniero/perfil_ingeniero.html', context)

@login_required
def perfil_cliente(request):
    user = request.user
    perfil, created = PerfilUsuario.objects.get_or_create(user=user)
    cotizaciones = Cotizacion.objects.filter(cliente=user).order_by('-fecha')[:5]
    
    context = {
        'user': user,
        'cotizaciones': cotizaciones,
        'perfil': perfil,
        'rol': 'Cliente',
        'avatar_url': perfil.avatar.url if perfil.avatar else '/static/images/default-avatar.png'
    }
    
    return render(request, 'ingeniero/perfil_cliente.html', context)

@login_required
def perfil_ingeniero(request):
    user = request.user
    # Get or create the profile if it doesn't exist
    perfil, created = PerfilUsuario.objects.get_or_create(user=user)
    cotizaciones = Cotizacion.objects.filter(ingeniero=user).order_by('-fecha')[:5]
    
    return render(request, 'ingeniero/perfil_cliente.html', {
        'user': user,
        'cotizaciones': cotizaciones,
        'perfil': perfil  # Use the profile we just got/created
    })

@login_required
def perfil_cliente(request):
    user = request.user
    # Get or create the profile if it doesn't exist
    perfil, created = PerfilUsuario.objects.get_or_create(user=user)
    cotizaciones = Cotizacion.objects.filter(ingeniero=user).order_by('-fecha')[:5]
    
    return render(request, 'ingeniero/perfil_ingeniero.html', {
        'user': user,
        'cotizaciones': cotizaciones,
        'perfil': perfil  # Use the profile we just got/created
    })

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Cotizacion

@login_required
def rechazar_cotizacion(request, id):
    cotizacion = get_object_or_404(Cotizacion, id=id)
    
    
    if request.method == 'POST':
        motivo = request.POST.get('motivo', '').strip()
        
        if not motivo:
            messages.error(request, 'Debes ingresar un motivo de rechazo')
            return redirect('revisar_cotizacion', cotizacion_id=id)
        
        # Actualizar estado y motivo
        cotizacion.estado = 'rechazado'
        cotizacion.motivo_rechazo = motivo
        cotizacion.save()
        
        messages.success(request, 'Cotización rechazada correctamente')
        return redirect('revisar_cotizacion', cotizacion_id=id)
    
    # Si no es POST, redirigir a revisar cotización
    return redirect('revisar_cotizacion', cotizacion_id=id)

def perfil_view(request):
    perfil = request.user.perfil  
    

    if request.user.groups.filter(name='CLIENTE').exists():
        cotizaciones = Cotizacion.objects.filter(cliente=request.user)
    else:
        cotizaciones = Cotizacion.objects.all().order_by('-fecha')[:5]  \
    
    context = {
        'perfil': perfil,
        'cotizaciones': cotizaciones,
    }
    return render(request, 'ingeniero/perfil_cliente.html', context)
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Cotizacion, CotizacionItem

from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Cotizacion, CotizacionItem
from num2words import num2words

@login_required
def detalle_cotizacion(request, id):
    cotizacion = get_object_or_404(
        Cotizacion.objects.select_related('cliente', 'ingeniero')
                         .prefetch_related('items'),
        id=id
    )
    total_en_palabras = num2words(cotizacion.total, lang='es')
    
    if request.method == 'POST' and 'aprobar_cotizacion' in request.POST:
            cotizacion.estado = 'aprobado'
            cotizacion.save()
            messages.success(request, 'Cotización aprobada exitosamente')
            return redirect('detalle_cotizacion', id=id)


    if not (request.user == cotizacion.cliente or 
            request.user == cotizacion.ingeniero or 
            request.user.is_superuser):
        messages.error(request, 'No tienes permiso para ver esta cotización')
        return redirect('home')
    
    
    context = {
        'total_en_palabras': total_en_palabras,
        'cotizacion': cotizacion,
        'items': cotizacion.items.all(),
        'puede_editar': request.user == cotizacion.cliente and cotizacion.estado == 'BORRADOR',
        'puede_revisar': request.user == cotizacion.ingeniero and cotizacion.estado == 'pendiente',
    }
    
    return render(request, 'cotizacion/detalle.html', context)
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from io import BytesIO

def generar_contrato_pdf(request, cotizacion_id):
    cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id)
    items = cotizacion.items.all()
    
    # Renderiza el template HTML
    html_string = render_to_string('pdf/contrato_pdf.html', {
        'cotizacion': cotizacion,
        'items': items,
        'fecha_actual': timezone.now().strftime("%d/%m/%Y")
    })
    
    # Crea el PDF
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf_file = html.write_pdf()
    
    # Configura la respuesta
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Contrato_{cotizacion.id}.pdf"'
    return response

def Mantenimiento(request):
    return render(request, 'servicios/Mantenimiento.html')

def Construccion(request):
    return render(request, 'servicios/Construccion.html')
def Proyectos(request):
    return render(request, 'servicios/Proyectos.html')

@login_required
def generar_reporte_cotizaciones(request):
    # Obtener cotizaciones pendientes
    cotizaciones = Cotizacion.objects.filter(estado='BORRADOR').order_by('-fecha')

    # Preparar contexto
    context = {
        'cotizaciones': cotizaciones,
        'fecha_reporte': timezone.now().strftime("%d/%m/%Y %H:%M"),
        'usuario': request.user.get_full_name() or request.user.username,
    }

    # Renderizar HTML desde el template
    html_string = render_to_string('pdf/cotizaciones_pdf.html', context)

    # Convertir HTML a PDF con WeasyPrint
    pdf_file = BytesIO()
    HTML(string=html_string).write_pdf(pdf_file)

    # Devolver PDF como respuesta HTTP
    response = HttpResponse(pdf_file.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Cotizaciones_Pendientes.pdf"'
    return response
    
    if pisa_status.err:
        return HttpResponse('Error al generar PDF')
    return response

# views.py
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings

# views.py
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages

@login_required
def enviar_cotizacion(request, cotizacion_id):
    cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id, ingeniero=request.user)
    
    # Contexto sin referencia a proyecto
    context = {
        'cotizacion': cotizacion,
        'cliente_nombre': cotizacion.cliente.get_full_name() or cotizacion.cliente.username
    }
    
    email_subject = f'Cotización #{cotizacion.id} - YGV Construcciones'
    email_body = render_to_string('ingeniero/email_cotizacion.html', context)
    
    try:
        send_mail(
            email_subject,
            '',  # Mensaje vacío porque usamos html_message
            settings.DEFAULT_FROM_EMAIL,
            [cotizacion.cliente.email],
            html_message=email_body,
            fail_silently=False,
        )
        messages.success(request, f'Cotización enviada a {cotizacion.cliente.email}')
    except Exception as e:
        messages.error(request, f'Error al enviar: {str(e)}')
    
    return redirect('home')


import base64
from django.shortcuts import get_object_or_404, redirect
from .models import Cotizacion


def guardar_firma(request, cotizacion_id):
    if request.method == 'POST':
        cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id)
        firma_data = request.POST.get('firma')
        if firma_data:
            cotizacion.firma_base64 = firma_data
            cotizacion.estado = 'aprobado'
            cotizacion.save()
        return redirect('detalle_cotizacion', id=cotizacion.id)
